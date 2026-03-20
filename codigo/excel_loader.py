from __future__ import annotations

import os
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook

from models import Circunscripcion, EleccionCongreso2023, Partido, ResultadoPartido


class ExcelStructureError(Exception):
    pass


class ElectionDataLoader:
    def __init__(self, excel_path: str) -> None:
        self.excel_path = excel_path
        self.column_aliases = {
            "circunscripcion_codigo": [
                "codprovincia",
                "cod_provincia",
                "codigo_provincia",
                "cod circunscripcion",
                "codcircunscripcion",
                "codigo circunscripcion",
                "circunscripcion codigo",
                "provincia codigo",
                "provincia codigo ine",
                "cp",
                "cpro",
                "nprov",
            ],
            "circunscripcion_nombre": [
                "provincia",
                "nombre provincia",
                "circunscripcion",
                "nombre circunscripcion",
                "provincia circunscripcion",
                "nprovincia",
                "nombre prov",
                "nombre de provincia",
            ],
            "comunidad_autonoma": [
                "comunidad autonoma",
                "autonomia",
                "ccaa",
                "nombre comunidad autonoma",
                "comunidad",
                "nombre comunidad",
                "cca",
            ],
            "partido_codigo": [
                "cod candidatura",
                "cod candidatura acumulado",
                "codpartido",
                "codigo partido",
                "codigo candidatura",
                "sigla candidatura",
                "cod cand",
                "candidatura codigo",
                "ccand",
                "cand",
            ],
            "partido_nombre": [
                "denominacion candidatura",
                "candidatura",
                "nombre candidatura",
                "partido",
                "nombre partido",
                "denominacion",
                "nombre candidatura completa",
            ],
            "partido_sigla": [
                "siglas candidatura",
                "sigla",
                "siglas",
                "abreviatura candidatura",
                "siglas partido",
            ],
            "votos": [
                "votos",
                "num votos",
                "votos candidatura",
                "votos obtenidos",
                "votoscand",
                "vot",
            ],
            "escanos_oficiales_partido": [
                "diputados",
                "escanos",
                "escanos partido",
                "diputados electos",
                "electos",
                "dip",
            ],
            "escanos_circunscripcion": [
                "numero diputados",
                "diputados a elegir",
                "escanos circunscripcion",
                "diputados circunscripcion",
                "num diputados",
                "diputados provincia",
                "escanos provincia",
                "dipu",
                "np" ,
            ],
            "votos_totales_candidaturas": [
                "votos a candidaturas",
                "votos candidaturas",
                "total votos candidaturas",
                "votos validos",
                "votcands",
                "votos validos candidaturas",
            ],
        }
        self.required_fields = [
            "circunscripcion_codigo",
            "circunscripcion_nombre",
            "partido_codigo",
            "partido_nombre",
            "votos",
            "escanos_circunscripcion",
        ]

    def load_election(self) -> Tuple[EleccionCongreso2023, List[str]]:
        if not os.path.exists(self.excel_path):
            raise FileNotFoundError(
                "No se encontro el archivo Excel en la ruta: {0}".format(self.excel_path)
            )

        rows, header_row_index, sheet_name = self._read_candidate_rows()
        if len(rows) == 0:
            raise ExcelStructureError("No se encontraron filas de datos utilizables en el Excel.")

        headers = list(rows[0].keys())
        column_mapping = self._resolve_column_mapping(headers)
        normalized_rows = self._prepare_rows(rows, column_mapping)
        if len(normalized_rows) == 0:
            raise ExcelStructureError(
                "Se localizaron cabeceras en la hoja {0} (fila {1}), pero ninguna fila de datos cumplio las validaciones minimas.".format(
                    sheet_name, header_row_index + 1
                )
            )

        election = EleccionCongreso2023(
            nombre="Elecciones generales al Congreso 2023",
            archivo_origen=self.excel_path,
            metadatos_columnas=column_mapping,
        )
        messages: List[str] = []
        messages.append(
            "CONFIRMACION: Se selecciono la hoja '{0}' y la fila de cabeceras {1}.".format(
                sheet_name, header_row_index + 1
            )
        )

        for row in normalized_rows:
            circunscripcion = self._get_or_create_circunscripcion(election, row)
            partido = self._build_partido(row)
            messages.append(election.registrar_partido(partido))
            resultado = ResultadoPartido(
                partido=partido,
                votos=int(row["votos"]),
                escanos_oficiales=int(row["escanos_oficiales_partido"]),
            )
            messages.append(circunscripcion.agregar_resultado(resultado))

        return election, messages

    def _read_candidate_rows(self) -> Tuple[List[Dict[str, object]], int, str]:
        workbook = load_workbook(self.excel_path, read_only=True, data_only=True)
        if len(workbook.sheetnames) == 0:
            raise ExcelStructureError("El libro Excel no contiene hojas.")

        best_candidate_rows: List[Dict[str, object]] = []
        best_candidate_header_index = -1
        best_candidate_sheet_name = workbook.sheetnames[0]
        best_candidate_score = -1
        best_candidate_headers: List[str] = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sampled_rows = self._read_non_empty_rows(sheet, maximum_rows=80)
            if len(sampled_rows) == 0:
                continue

            header_index, headers, score = self._find_header_row(sampled_rows)
            if header_index < 0:
                continue

            candidate_rows = self._build_rows_from_header(sheet, header_index, headers)
            if score > best_candidate_score or (
                score == best_candidate_score and len(candidate_rows) > len(best_candidate_rows)
            ):
                best_candidate_rows = candidate_rows
                best_candidate_header_index = header_index
                best_candidate_sheet_name = sheet_name
                best_candidate_score = score
                best_candidate_headers = headers

        workbook.close()

        if best_candidate_header_index < 0:
            raise ExcelStructureError(
                "No se pudo localizar una fila de cabeceras valida en ninguna hoja del Excel."
            )

        if best_candidate_score < len(self.required_fields):
            raise ExcelStructureError(
                "No se pudieron identificar las columnas obligatorias. Columnas candidatas encontradas: {0}".format(
                    ", ".join(best_candidate_headers)
                )
            )

        return best_candidate_rows, best_candidate_header_index, best_candidate_sheet_name

    def _read_non_empty_rows(self, sheet, maximum_rows: int) -> List[List[object]]:
        rows: List[List[object]] = []
        row_counter = 0
        for data_row in sheet.iter_rows(values_only=True):
            normalized_row: List[object] = []
            has_content = False
            for value in data_row:
                normalized_row.append(value)
                if value is not None and str(value).strip() != "":
                    has_content = True
            if has_content:
                rows.append(normalized_row)
            row_counter = row_counter + 1
            if row_counter >= maximum_rows:
                break
        return rows

    def _find_header_row(self, sampled_rows: List[List[object]]) -> Tuple[int, List[str], int]:
        best_index = -1
        best_headers: List[str] = []
        best_score = -1
        row_index = 0
        while row_index < len(sampled_rows):
            row = sampled_rows[row_index]
            headers = self._row_to_headers(row)
            mapping = self._resolve_column_mapping(headers, raise_error=False)
            score = len(mapping)
            if score > best_score:
                best_index = row_index
                best_headers = headers
                best_score = score
            row_index = row_index + 1
        return best_index, best_headers, best_score

    def _row_to_headers(self, row: List[object]) -> List[str]:
        headers: List[str] = []
        index = 0
        while index < len(row):
            value = row[index]
            header = self._as_text(value)
            if header == "":
                header = "columna_{0}".format(index + 1)
            headers.append(header)
            index = index + 1
        return headers

    def _build_rows_from_header(self, sheet, header_index: int, headers: List[str]) -> List[Dict[str, object]]:
        rows: List[Dict[str, object]] = []
        current_index = 0
        for data_row in sheet.iter_rows(values_only=True):
            if current_index <= header_index:
                current_index = current_index + 1
                continue
            row_dictionary: Dict[str, object] = {}
            is_empty = True
            column_index = 0
            while column_index < len(headers):
                header = headers[column_index]
                value = None
                if column_index < len(data_row):
                    value = data_row[column_index]
                row_dictionary[header] = value
                if value is not None and str(value).strip() != "":
                    is_empty = False
                column_index = column_index + 1
            if not is_empty:
                rows.append(row_dictionary)
            current_index = current_index + 1
        return rows

    def _normalize_text(self, value: object) -> str:
        if value is None:
            return ""
        text = str(value).strip().lower()
        replacements = {
            "á": "a",
            "é": "e",
            "í": "i",
            "ó": "o",
            "ú": "u",
            "ü": "u",
            "ñ": "n",
            "_": " ",
            "-": " ",
            ".": " ",
            "/": " ",
            "(": " ",
            ")": " ",
        }
        for source, target in replacements.items():
            text = text.replace(source, target)
        text = " ".join(text.split())
        return text

    def _resolve_column_mapping(
        self, headers: List[str], raise_error: bool = True
    ) -> Dict[str, str]:
        normalized_source_columns: Dict[str, str] = {}
        for column in headers:
            normalized_column = self._normalize_text(column)
            normalized_source_columns[normalized_column] = str(column)

        mapping: Dict[str, str] = {}
        for logical_name, aliases in self.column_aliases.items():
            resolved = self._search_column_by_alias(normalized_source_columns, aliases)
            if resolved is not None:
                mapping[logical_name] = resolved

        if raise_error:
            missing_fields: List[str] = []
            for field_name in self.required_fields:
                if field_name not in mapping:
                    missing_fields.append(field_name)

            if len(missing_fields) > 0:
                raise ExcelStructureError(
                    "No se pudieron identificar las columnas obligatorias: {0}. Columnas encontradas: {1}".format(
                        ", ".join(missing_fields), ", ".join(headers)
                    )
                )
        return mapping

    def _search_column_by_alias(
        self, normalized_source_columns: Dict[str, str], aliases: List[str]
    ) -> Optional[str]:
        for alias in aliases:
            normalized_alias = self._normalize_text(alias)
            if normalized_alias in normalized_source_columns:
                return normalized_source_columns[normalized_alias]
        for normalized_name, original_name in normalized_source_columns.items():
            for alias in aliases:
                normalized_alias = self._normalize_text(alias)
                if normalized_alias != "" and normalized_alias in normalized_name:
                    return original_name
                if normalized_name != "" and normalized_name in normalized_alias:
                    return original_name
        return None

    def _prepare_rows(
        self, rows: List[Dict[str, object]], column_mapping: Dict[str, str]
    ) -> List[Dict[str, object]]:
        normalized_rows: List[Dict[str, object]] = []
        for original_row in rows:
            normalized_row: Dict[str, object] = {}
            for logical_name, source_name in column_mapping.items():
                normalized_row[logical_name] = original_row.get(source_name)

            normalized_row["comunidad_autonoma"] = self._as_text(
                normalized_row.get("comunidad_autonoma", "")
            )
            normalized_row["partido_sigla"] = self._as_text(normalized_row.get("partido_sigla", ""))
            normalized_row["circunscripcion_codigo"] = self._normalize_numeric_code(
                normalized_row.get("circunscripcion_codigo", "")
            )
            normalized_row["partido_codigo"] = self._normalize_numeric_code(
                normalized_row.get("partido_codigo", "")
            )
            normalized_row["circunscripcion_nombre"] = self._as_text(
                normalized_row.get("circunscripcion_nombre", "")
            )
            normalized_row["partido_nombre"] = self._as_text(normalized_row.get("partido_nombre", ""))
            normalized_row["votos"] = self._as_integer(normalized_row.get("votos"))
            normalized_row["escanos_circunscripcion"] = self._as_integer(
                normalized_row.get("escanos_circunscripcion")
            )
            normalized_row["escanos_oficiales_partido"] = self._as_integer(
                normalized_row.get("escanos_oficiales_partido"), default_value=0
            )
            normalized_row["votos_totales_candidaturas"] = self._as_optional_integer(
                normalized_row.get("votos_totales_candidaturas")
            )

            if normalized_row["votos"] is None or normalized_row["escanos_circunscripcion"] is None:
                continue
            if int(normalized_row["votos"]) <= 0:
                continue
            if normalized_row["circunscripcion_nombre"] == "" or normalized_row["partido_nombre"] == "":
                continue
            if normalized_row["circunscripcion_codigo"] == "" or normalized_row["partido_codigo"] == "":
                continue
            normalized_rows.append(normalized_row)
        return normalized_rows

    def _normalize_numeric_code(self, value: object) -> str:
        if value is None:
            return ""
        text = str(value).strip()
        if text == "":
            return ""
        try:
            numeric_value = float(text)
            if numeric_value.is_integer():
                return str(int(numeric_value)).zfill(2)
        except (TypeError, ValueError):
            pass
        return text

    def _as_text(self, value: object) -> str:
        if value is None:
            return ""
        return str(value).strip()

    def _as_integer(self, value: object, default_value: Optional[int] = None) -> Optional[int]:
        if value is None or str(value).strip() == "":
            return default_value
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return default_value

    def _as_optional_integer(self, value: object) -> Optional[int]:
        return self._as_integer(value, default_value=None)

    def _get_or_create_circunscripcion(
        self, election: EleccionCongreso2023, row: Dict[str, object]
    ) -> Circunscripcion:
        codigo = str(row["circunscripcion_codigo"])
        if codigo in election.circunscripciones:
            circunscripcion_existente = election.circunscripciones[codigo]
            if circunscripcion_existente.votos_totales_candidaturas_oficiales is None:
                circunscripcion_existente.votos_totales_candidaturas_oficiales = row.get(
                    "votos_totales_candidaturas"
                )
            return circunscripcion_existente

        circunscripcion = Circunscripcion(
            codigo=codigo,
            nombre=str(row["circunscripcion_nombre"]),
            provincia=str(row["circunscripcion_nombre"]),
            comunidad_autonoma=str(row.get("comunidad_autonoma", "")),
            escanos_oficiales_totales=int(row["escanos_circunscripcion"]),
            votos_totales_candidaturas_oficiales=row.get("votos_totales_candidaturas"),
        )
        election.registrar_circunscripcion(circunscripcion)
        return circunscripcion

    def _build_partido(self, row: Dict[str, object]) -> Partido:
        return Partido(
            codigo=str(row["partido_codigo"]),
            nombre=str(row["partido_nombre"]),
            sigla=str(row.get("partido_sigla", "")),
        )
